# Great-looking Tkinter app to clean XLSX files by keeping only a few columns.
# Dependencies: Python 3.x, Tkinter (bundled), openpyxl (installed in the local venv).

from __future__ import annotations

import re
import threading
from pathlib import Path
from tkinter import Canvas, StringVar, Tk, filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

PRIMARY_BG = "#0b1224"
GRADIENT_START = "#0b1224"
GRADIENT_END = "#10375f"
CARD_BG = "#0f213a"
CARD_BORDER = "#1e3a5f"
ACCENT = "#22d3ee"
TEXT_MAIN = "#e5f0ff"
TEXT_MUTED = "#b8c5d6"
TEXT_ACCENT = "#7dd3fc"

DESIRED_COLUMNS = [
    "order_id",
    "product_name",
    "quantity",
    "total_quantity",
]

# Normalize headers to match user input variants and dynamic numbered fields.
HEADER_ALIASES = {
    "orderid": "order_id",
    "order_id": "order_id",
    "order id": "order_id",
    "id": "order_id",
    "qty": "quantity",
    "quantity": "quantity",
    "totalquantity": "total_quantity",
    "total_quantity": "total_quantity",
    "total quantity": "total_quantity",
    "total_qty": "total_quantity",
}


class GradientFrame(ttk.Frame):
    """Simple gradient background using a canvas behind the content."""

    def __init__(self, master: Tk, color1: str, color2: str, **kwargs):
        super().__init__(master, **kwargs)
        self._color1 = color1
        self._color2 = color2
        self._canvas = Canvas(self, highlightthickness=0, bd=0)
        self._canvas.pack(fill="both", expand=True)
        self._canvas.bind("<Configure>", self._draw)
        self._content = ttk.Frame(self._canvas, style="Card.TFrame")
        self._content_window = self._canvas.create_window(
            (0, 0), window=self._content, anchor="nw"
        )

    def content(self) -> ttk.Frame:
        return self._content

    def _draw(self, event=None) -> None:
        self._canvas.delete("gradient")
        width = self._canvas.winfo_width()
        height = self._canvas.winfo_height()
        self._canvas.itemconfig(
            self._content_window, width=width, height=height
        )
        # Vertical gradient by stacking rectangles.
        limit = max(height, 1)
        for i in range(limit):
            ratio = i / limit
            r1, g1, b1 = self._canvas.winfo_rgb(self._color1)
            r2, g2, b2 = self._canvas.winfo_rgb(self._color2)
            r = int(r1 + (r2 - r1) * ratio)
            g = int(g1 + (g2 - g1) * ratio)
            b = int(b1 + (b2 - b1) * ratio)
            color = f"#{r:04x}{g:04x}{b:04x}"
            self._canvas.create_line(
                0, i, width, i, tags=("gradient",), fill=color
            )
        self._canvas.lower("gradient")


class ExcelCleanerUI:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("XLSX Cleaner • Keep what matters")
        self.root.geometry("720x520")
        self.root.minsize(640, 460)
        self.root.configure(bg=PRIMARY_BG)
        self.file_var = StringVar()
        self.status_var = StringVar(value="Select an .xlsx file to start.")

        self._setup_styles()
        gradient = GradientFrame(root, GRADIENT_START, GRADIENT_END)
        gradient.pack(fill="both", expand=True)
        container = gradient.content()
        container.columnconfigure(0, weight=1)

        header = ttk.Label(
            container,
            text="XLSX Cleaner",
            style="Title.TLabel",
        )
        subtitle = ttk.Label(
            container,
            text="Keep only order id, product names, quantities, and total quantity.\nGroups rows by product name and sums total quantity per product.",
            style="Body.TLabel",
            anchor="center",
            justify="center",
        )
        header.grid(row=0, column=0, sticky="n", pady=(36, 8))
        subtitle.grid(row=1, column=0, sticky="n", padx=28, pady=(0, 18))

        card = ttk.Frame(container, style="Glass.TFrame", padding=22)
        card.grid(row=2, column=0, sticky="n", padx=32, pady=12)
        for i in range(2):
            card.columnconfigure(i, weight=1)

        path_label = ttk.Label(
            card,
            text="Choose your Excel (.xlsx) file",
            style="Caption.TLabel",
        )
        path_entry = ttk.Entry(
            card,
            textvariable=self.file_var,
            font=("Segoe UI", 11),
            justify="center",
        )
        browse_btn = ttk.Button(
            card,
            text="Browse",
            style="TButton",
            command=self.choose_file,
        )
        path_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        path_entry.grid(
            row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10)
        )
        browse_btn.grid(row=2, column=0, columnspan=2, sticky="ew")

        action_btn = ttk.Button(
            card,
            text="Clean & Save",
            style="Accent.TButton",
            command=self._run_clean_async,
        )
        action_btn.grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(16, 0)
        )

        self.status_label = ttk.Label(
            container,
            textvariable=self.status_var,
            style="Status.TLabel",
            anchor="center",
        )
        self.status_label.grid(row=3, column=0, sticky="ew", padx=28, pady=14)

        footer = ttk.Label(
            container,
            text="Made for quick clean-ups • Saves next to this script",
            style="Caption.TLabel",
        )
        footer.grid(row=4, column=0, sticky="n", pady=(0, 18))

    def _setup_styles(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            ".",
            background=PRIMARY_BG,
            foreground=TEXT_MAIN,
            font=("Segoe UI", 11),
        )
        style.configure(
            "Title.TLabel",
            font=("Segoe UI", 26, "bold"),
            foreground="#e7f8ff",
        )
        style.configure(
            "Body.TLabel", font=("Segoe UI", 12), foreground=TEXT_MUTED
        )
        style.configure(
            "Status.TLabel",
            font=("Segoe UI", 11, "italic"),
            foreground=TEXT_ACCENT,
            anchor="center",
        )
        style.configure(
            "Caption.TLabel",
            font=("Segoe UI", 10, "bold"),
            foreground=TEXT_ACCENT,
            anchor="center",
        )

        style.configure(
            "Card.TFrame",
            background=PRIMARY_BG,
            borderwidth=0,
        )
        style.configure(
            "Glass.TFrame",
            background=CARD_BG,
            borderwidth=1.2,
            relief="solid",
            bordercolor=CARD_BORDER,
        )
        style.configure(
            "TEntry",
            fieldbackground="#0c1c33",
            foreground=TEXT_MAIN,
            bordercolor=CARD_BORDER,
            insertcolor=TEXT_MAIN,
        )
        style.map("TEntry", bordercolor=[("focus", ACCENT)])

        style.configure(
            "Accent.TButton",
            background=ACCENT,
            foreground=PRIMARY_BG,
            font=("Segoe UI", 12, "bold"),
            padding=10,
            borderwidth=0,
            focusthickness=3,
            focuscolor=ACCENT,
        )
        style.map(
            "Accent.TButton",
            background=[("active", "#4de5ff")],
            foreground=[("disabled", "#94a3b8")],
        )
        style.configure(
            "TButton",
            font=("Segoe UI", 11, "bold"),
            padding=9,
            background="#1a365d",
            foreground=TEXT_MAIN,
            borderwidth=1,
            relief="solid",
            bordercolor=CARD_BORDER,
        )
        style.map(
            "TButton",
            background=[("active", "#254572")],
            bordercolor=[("focus", ACCENT)],
        )

    def choose_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Choose an Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.file_var.set(path)
            self.status_var.set("Ready to clean.")

    def _run_clean_async(self) -> None:
        path = self.file_var.get().strip()
        if not path:
            messagebox.showwarning(
                "No file selected", "Please choose an .xlsx file first."
            )
            return
        self.status_var.set("Working…")
        threading.Thread(
            target=self._clean_file, args=(Path(path),), daemon=True
        ).start()

    def _clean_file(self, path: Path) -> None:
        try:
            output = clean_excel(path)
        except Exception as exc:  # pylint: disable=broad-except
            self.status_var.set("Something went wrong. Check the message.")
            messagebox.showerror("Error while cleaning", str(exc))
            return
        self.status_var.set(f"Saved cleaned file: {output.name}")
        messagebox.showinfo("Done", f"Cleaned file saved as:\n{output}")


def normalize_header(name: str) -> str:
    key = re.sub(r"[\s_-]+", "", (name or "").strip().lower())
    return HEADER_ALIASES.get(key, key)


def normalize_product_name(name) -> str:
    if name is None:
        return ""
    text = re.sub(r"\s+", " ", str(name).strip().lower())
    return text


def clean_excel(file_path: Path) -> Path:
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path)
    sheet = wb.active

    header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1)))
    product_name_cols = {}
    quantity_cols = {}
    order_col = None
    for idx, cell in enumerate(header_cells):
        key = normalize_header(cell.value or "")
        if key == "order_id":
            order_col = idx
            continue
        name_match = re.match(r"productname(\d*)", key)
        qty_match = re.match(r"(productquantity|quantity)(\d*)", key)
        if name_match:
            num = int(name_match.group(1) or "1")
            product_name_cols[num] = idx
        if qty_match:
            num = int(qty_match.group(2) or "1")
            quantity_cols[num] = idx

    if order_col is None:
        raise ValueError("Missing required column: id / order_id")
    if not product_name_cols and not quantity_cols:
        raise ValueError("No productName / productQuantity columns found.")

    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Cleaned"
    new_sheet.append([col.replace("_", " ").title() for col in DESIRED_COLUMNS])

    def parse_quantity(value):
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def format_number(num):
        if num is None:
            return None
        if isinstance(num, float) and num.is_integer():
            return int(num)
        return num

    extracted_rows = []
    product_totals = {}
    product_has_numeric = {}
    product_display_names = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        order_id = row[order_col] if order_col < len(row) else None
        if order_id is None:
            continue
        for num in sorted(set(product_name_cols) | set(quantity_cols)):
            name_idx = product_name_cols.get(num)
            qty_idx = quantity_cols.get(num)
            product_name = (
                row[name_idx]
                if name_idx is not None and name_idx < len(row)
                else None
            )
            qty_raw = (
                row[qty_idx]
                if qty_idx is not None and qty_idx < len(row)
                else None
            )
            qty_value = parse_quantity(qty_raw)
            if product_name is None and qty_raw is None:
                continue

            product_key = normalize_product_name(product_name)
            if product_key and product_key not in product_display_names:
                product_display_names[product_key] = product_name
            if qty_value is not None:
                product_totals[product_key] = (
                    product_totals.get(product_key, 0.0) + qty_value
                )
                product_has_numeric[product_key] = True
            else:
                product_totals.setdefault(product_key, 0.0)

            extracted_rows.append(
                {
                    "product_key": product_key,
                    "order_id": order_id,
                    "product_name": product_name,
                    "quantity_display": format_number(qty_value)
                    if qty_value is not None
                    else qty_raw,
                    "row_index": len(extracted_rows),
                }
            )

    def sort_key(item):
        product_key = item["product_key"] or "\uffff"
        return (product_key, str(item["order_id"]), item["row_index"])

    extracted_rows.sort(key=sort_key)

    def total_for_product_key(product_key: str):
        if product_has_numeric.get(product_key):
            return format_number(product_totals.get(product_key))
        return None

    current_product_key = None
    group_start_row = None
    excel_row = 2  # header is on row 1

    def finalize_group(product_key: str, start_row: int, end_row: int) -> None:
        total_display = total_for_product_key(product_key)
        cell = new_sheet.cell(row=start_row, column=4)
        cell.value = total_display
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if end_row > start_row:
            new_sheet.merge_cells(
                start_row=start_row,
                start_column=4,
                end_row=end_row,
                end_column=4,
            )

    for item in extracted_rows:
        product_key = item["product_key"]
        if current_product_key is None:
            current_product_key = product_key
            group_start_row = excel_row
        elif product_key != current_product_key:
            finalize_group(current_product_key, group_start_row, excel_row - 1)
            current_product_key = product_key
            group_start_row = excel_row

        new_sheet.append(
            [
                item["order_id"],
                item["product_name"],
                item["quantity_display"],
                None,
            ]
        )
        excel_row += 1

    if current_product_key is not None and group_start_row is not None:
        finalize_group(current_product_key, group_start_row, excel_row - 1)

    def autosize_columns(
        sheet, *, min_width: int = 8, max_width: int = 60, padding: int = 2
    ) -> None:
        for col_idx in range(1, len(DESIRED_COLUMNS) + 1):
            max_len = 0
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                value = row[0].value
                if value is None:
                    continue
                text = str(value).replace("\n", " ").strip()
                if len(text) > max_len:
                    max_len = len(text)
            width = max(min_width, min(max_width, max_len + padding))
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    autosize_columns(new_sheet, max_width=80)

    summary_sheet = new_wb.create_sheet(title="Summary")
    summary_sheet.append(["Output Summary"])
    summary_sheet.append(
        [
            "Total quantity (numeric)",
            format_number(sum(product_totals.values())),
        ]
    )
    summary_sheet.append(
        ["Unique products", len([k for k in product_totals if k])]
    )
    summary_sheet.append([])
    summary_sheet.append(["Product name", "Total quantity (numeric)"])

    summary_rows = []
    for key, total in product_totals.items():
        if not key:
            continue
        if not product_has_numeric.get(key):
            continue
        display_name = product_display_names.get(key, key)
        summary_rows.append((display_name, total))
    summary_rows.sort(key=lambda x: x[1], reverse=True)
    for name, total in summary_rows:
        summary_sheet.append([name, format_number(total)])

    if summary_rows:
        best_name, best_total = summary_rows[0]
        summary_sheet.append([])
        summary_sheet.append(["Best-selling product", best_name])
        summary_sheet.append(
            ["Sales amount (numeric)", format_number(best_total)]
        )

    summary_sheet.column_dimensions["A"].width = 40
    summary_sheet.column_dimensions["B"].width = 22

    output_name = f"{file_path.stem}_cleaned.xlsx"
    output_path = Path.cwd() / output_name
    new_wb.save(output_path)
    return output_path


def main() -> None:
    root = Tk()
    app = ExcelCleanerUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
